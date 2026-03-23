import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from load_and_validate import load_all_data
from validate_data import run_validation_checks
from reconcile_trades import reconcile_executed_vs_booked

# =========================
# PATH SETUP
# =========================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

EXCEPTIONS_REPORT_FILE = os.path.join(OUTPUT_DIR, "exceptions_report.xlsx")
DAILY_SUMMARY_FILE = os.path.join(OUTPUT_DIR, "daily_summary.csv")

# =========================
# CONFIG
# =========================
SEVERITY_ORDER = {
    "High": 1,
    "Medium": 2,
    "Low": 3,
}

VALIDATION_ACTIONS = {
    "duplicate_account_id": "Review the accounts reference and remove or merge the duplicate account record.",
    "inactive_account_used": "Review the account status and confirm whether the trade should be booked to an active account.",
}


# =========================
# HELPERS
# =========================
def get_validation_recommended_action(issue_type):
    """
    Return a consistent follow-up action for each validation issue type.
    """
    if issue_type.startswith("missing_"):
        return "Review the source row and populate the required field before rerunning checks."

    return VALIDATION_ACTIONS.get(
        issue_type,
        "Review the source row and correct the data-quality issue before rerunning checks.",
    )


def standardise_validation_issues(validation_df):
    """
    Convert data-check items into the same structure as reconciliation breaks.
    """
    if validation_df.empty:
        return pd.DataFrame(columns=[
            "trade_id",
            "source_file",
            "exception_type",
            "severity",
            "executed_value",
            "booked_value",
            "recommended_action",
            "status",
        ])

    validation_standardised = validation_df.copy()

    validation_standardised["source_file"] = validation_standardised["file_name"]
    validation_standardised["exception_type"] = validation_standardised["issue_type"]
    validation_standardised["executed_value"] = ""
    validation_standardised["booked_value"] = ""
    validation_standardised["recommended_action"] = validation_standardised["issue_type"].apply(
        get_validation_recommended_action
    )
    validation_standardised["status"] = "Open"

    validation_standardised = validation_standardised[[
        "trade_id",
        "source_file",
        "exception_type",
        "severity",
        "executed_value",
        "booked_value",
        "recommended_action",
        "status",
    ]]

    return validation_standardised


def standardise_reconciliation_issues(reconciliation_df):
    """
    Ensure reconciliation breaks are in the final reporting structure.
    """
    if reconciliation_df.empty:
        return pd.DataFrame(columns=[
            "trade_id",
            "source_file",
            "exception_type",
            "severity",
            "executed_value",
            "booked_value",
            "recommended_action",
            "status",
        ])

    reconciliation_standardised = reconciliation_df.copy()

    reconciliation_standardised = reconciliation_standardised[[
        "trade_id",
        "source_file",
        "exception_type",
        "severity",
        "executed_value",
        "booked_value",
        "recommended_action",
        "status",
    ]]

    return reconciliation_standardised


def combine_all_issues(validation_df, reconciliation_df):
    """
    Combine data-check items and reconciliation breaks into one final table.
    """
    validation_standardised = standardise_validation_issues(validation_df)
    reconciliation_standardised = standardise_reconciliation_issues(reconciliation_df)

    all_issues_df = pd.concat(
        [validation_standardised, reconciliation_standardised],
        ignore_index=True
    )

    if all_issues_df.empty:
        return pd.DataFrame(columns=[
            "trade_id",
            "source_file",
            "exception_type",
            "severity",
            "executed_value",
            "booked_value",
            "recommended_action",
            "status",
        ])

    all_issues_df["severity_rank"] = all_issues_df["severity"].map(SEVERITY_ORDER)

    all_issues_df = all_issues_df.sort_values(
        by=["severity_rank", "exception_type", "trade_id"],
        ascending=[True, True, True]
    )

    all_issues_df = all_issues_df.drop(columns=["severity_rank"]).reset_index(drop=True)

    return all_issues_df


def calculate_total_matched_trades(executed_df, booked_df):
    """
    Count unique trade_ids present in both executed and booked files.
    """
    executed_ids = set(executed_df["trade_id"].dropna().astype(str))
    booked_ids = set(booked_df["trade_id"].dropna().astype(str))

    matched_ids = executed_ids.intersection(booked_ids)
    return len(matched_ids)


def build_daily_summary(executed_df, booked_df, all_issues_df):
    """
    Build a tidy daily summary table.
    """
    summary_rows = []

    total_executed_trades = len(executed_df)
    total_booked_trades = len(booked_df)
    total_matched_trades = calculate_total_matched_trades(executed_df, booked_df)
    total_exceptions = len(all_issues_df)

    # Trade volume / headline metrics
    summary_rows.append({
        "summary_type": "trade_volume",
        "summary_value": "total_executed_trades",
        "count": total_executed_trades,
    })
    summary_rows.append({
        "summary_type": "trade_volume",
        "summary_value": "total_booked_trades",
        "count": total_booked_trades,
    })
    summary_rows.append({
        "summary_type": "trade_volume",
        "summary_value": "total_matched_trades",
        "count": total_matched_trades,
    })
    summary_rows.append({
        "summary_type": "exceptions",
        "summary_value": "total_exceptions",
        "count": total_exceptions,
    })

    if not all_issues_df.empty:
        # By exception type
        exception_counts = all_issues_df["exception_type"].value_counts()
        for exception_type, count in exception_counts.items():
            summary_rows.append({
                "summary_type": "exception_type",
                "summary_value": exception_type,
                "count": count,
            })

        # By severity
        severity_counts = all_issues_df["severity"].value_counts()
        for severity, count in severity_counts.items():
            summary_rows.append({
                "summary_type": "severity",
                "summary_value": severity,
                "count": count,
            })

    summary_df = pd.DataFrame(summary_rows)
    return summary_df


def autosize_worksheet_columns(worksheet):
    """
    Auto-size worksheet columns based on content length.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def write_exceptions_report_excel(all_issues_df, executed_df, booked_df):
    """
    Write the final exception report to Excel with:
    - Exceptions sheet
    - Report_Info sheet
    - filters enabled
    - freeze top row
    - timestamp
    """
    generated_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_executed_trades = len(executed_df)
    total_booked_trades = len(booked_df)
    total_matched_trades = calculate_total_matched_trades(executed_df, booked_df)
    total_exceptions = len(all_issues_df)

    report_info_df = pd.DataFrame([
        {"field": "report_name", "value": "Trade Support Exception Report"},
        {"field": "generated_timestamp", "value": generated_timestamp},
        {"field": "total_executed_trades", "value": total_executed_trades},
        {"field": "total_booked_trades", "value": total_booked_trades},
        {"field": "total_matched_trades", "value": total_matched_trades},
        {"field": "total_exceptions", "value": total_exceptions},
    ])

    # Write using pandas first
    with pd.ExcelWriter(EXCEPTIONS_REPORT_FILE, engine="openpyxl") as writer:
        all_issues_df.to_excel(writer, sheet_name="Exceptions", index=False)
        report_info_df.to_excel(writer, sheet_name="Report_Info", index=False)

    # Re-open with openpyxl for formatting
    workbook = load_workbook(EXCEPTIONS_REPORT_FILE)

    # Exceptions sheet formatting
    exceptions_ws = workbook["Exceptions"]
    exceptions_ws.freeze_panes = "A2"

    if exceptions_ws.max_row > 1 and exceptions_ws.max_column > 1:
        exceptions_ws.auto_filter.ref = exceptions_ws.dimensions

    autosize_worksheet_columns(exceptions_ws)

    # Report info sheet formatting
    info_ws = workbook["Report_Info"]
    info_ws.freeze_panes = "A2"

    if info_ws.max_row > 1 and info_ws.max_column > 1:
        info_ws.auto_filter.ref = info_ws.dimensions

    autosize_worksheet_columns(info_ws)

    workbook.save(EXCEPTIONS_REPORT_FILE)


def save_daily_summary(summary_df):
    """
    Save daily summary to CSV.
    """
    summary_df.to_csv(DAILY_SUMMARY_FILE, index=False)


# =========================
# MAIN
# =========================
def main():
    try:
        # 1. Load all data
        executed_df, booked_df, accounts_df = load_all_data()

        # 2. Run validation
        validation_df = run_validation_checks(executed_df, booked_df, accounts_df)

        # 3. Run reconciliation
        reconciliation_df = reconcile_executed_vs_booked(executed_df, booked_df)

        # 4. Combine all issues
        all_issues_df = combine_all_issues(validation_df, reconciliation_df)

        # 5. Build daily summary
        summary_df = build_daily_summary(executed_df, booked_df, all_issues_df)

        # 6. Write outputs
        write_exceptions_report_excel(all_issues_df, executed_df, booked_df)
        save_daily_summary(summary_df)

        # 7. Console preview
        print("\nCombined exceptions:")
        print(all_issues_df)

        print("\nDaily summary:")
        print(summary_df)

        print(f"\nExceptions report saved to: {EXCEPTIONS_REPORT_FILE}")
        print(f"Daily summary saved to: {DAILY_SUMMARY_FILE}")

    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    main()
    
    